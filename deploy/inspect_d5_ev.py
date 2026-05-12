#!/usr/bin/env python3
import openpyxl
wb = openpyxl.load_workbook("/opt/whovoted/data/d5_ev_cumulative_final.xlsx")
sheet = wb.active
print(f"Rows: {sheet.max_row}, Cols: {sheet.max_column}")
for i in range(1, min(10, sheet.max_row+1)):
    row = [str(cell.value)[:40] for cell in sheet[i]]
    print(f"Row {i}: {row}")

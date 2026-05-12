#!/usr/bin/env python3
"""Download McAllen City Commission D5 election rosters and import."""
import urllib.request, os, json, sqlite3
from pathlib import Path

DATA_DIR = '/opt/whovoted/data'
DB_PATH = '/opt/whovoted/data/whovoted.db'

# Try various URL patterns for the rosters
BASE = 'https://www.mcallen.net/docs/default-source/cityelections/2026-Special-Election'
URLS_TO_TRY = [
    # Cumulative EV roster
    f'{BASE}/cumulative-report.xlsx',
    f'{BASE}/cumulative-report-april-28-2026.xlsx',
    f'{BASE}/ev-cumulative-roster.xlsx',
    f'{BASE}/april-28-2026.xlsx',
    # Election day roster
    f'{BASE}/election-day-roster.xlsx',
    f'{BASE}/may-2-2026.xlsx',
    f'{BASE}/election-day-may-2-2026.xlsx',
    f'{BASE}/ed-roster-may-2-2026.xlsx',
]

def try_download(url, outpath):
    """Try to download a file. Returns True if it's a real file (not HTML error)."""
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        resp = urllib.request.urlopen(req, timeout=30)
        data = resp.read()
        if data[:15].startswith(b'<!DOCTYPE') or data[:5] == b'<html':
            return False
        with open(outpath, 'wb') as f:
            f.write(data)
        return True
    except Exception as e:
        return False

def main():
    print("Trying to download D5 election rosters from McAllen city site...\n")
    
    found = []
    for url in URLS_TO_TRY:
        fname = url.split('/')[-1].split('?')[0]
        outpath = os.path.join(DATA_DIR, f'd5_{fname}')
        print(f"  {fname}...", end=' ')
        if try_download(url, outpath):
            size = os.path.getsize(outpath)
            print(f"✓ ({size/1024:.0f} KB)")
            found.append(outpath)
        else:
            print("✗")
    
    # We already have the mail ballot list — check it
    mail_path = os.path.join(DATA_DIR, 'd5_mail_ballots.xlsx')
    if os.path.exists(mail_path):
        print(f"\n  Already have: d5_mail_ballots.xlsx ({os.path.getsize(mail_path)/1024:.0f} KB)")
        found.append(mail_path)
    
    if not found:
        print("\nCould not download rosters automatically.")
        print("The links on the McAllen site may require JavaScript or have non-standard URLs.")
        print("\nManual option: download the Cumulative Report and Election Day Roster")
        print("from https://www.mcallen.net/departments/secretary/city-elections/special-election-2026")
        print("and place them in /opt/whovoted/data/")
        return
    
    # Try to read whatever we got
    print(f"\nFound {len(found)} files. Attempting to read...")
    try:
        import openpyxl
        for fpath in found:
            if not fpath.endswith('.xlsx'):
                continue
            print(f"\n  Reading {Path(fpath).name}...")
            wb = openpyxl.load_workbook(fpath)
            sheet = wb.active
            headers = [cell.value for cell in sheet[1]]
            print(f"    Headers: {headers}")
            print(f"    Rows: {sheet.max_row - 1}")
            # Show first 3 rows
            for row in sheet.iter_rows(min_row=2, max_row=min(4, sheet.max_row), values_only=True):
                print(f"    Sample: {row[:6]}")
    except ImportError:
        print("  openpyxl not installed — can't read xlsx files")
        print("  Install: pip install openpyxl")

if __name__ == '__main__':
    main()

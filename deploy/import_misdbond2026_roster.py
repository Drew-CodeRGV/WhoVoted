#!/usr/bin/env python3
"""
Import McAllen ISD Bond 2026 early voting rosters from Hidalgo County.

Downloads and processes cumulative early voting rosters.
URL: https://www.hidalgocounty.us/DocumentCenter/View/72488/EV-Roster-May-2-2026-Cumulative
"""

import requests
import sqlite3
import re
from datetime import datetime
import sys

DB_PATH = '/opt/whovoted/data/whovoted.db'
ELECTION_DATE = '2026-05-10'
ELECTION_NAME = 'McAllen ISD Bond 2026'

# Roster URL
ROSTER_URL = 'https://www.hidalgocounty.us/DocumentCenter/View/72534/EV-Roster-May-2-2026-Cumulative'

def download_roster(url):
    """Download roster file from URL."""
    print(f"Downloading roster from {url}...")
    response = requests.get(url, timeout=30)
    response.raise_for_status()
    return response.content

def parse_roster_pdf(content):
    """Parse roster file content to extract VUIDs."""
    # Try Excel first (XLSX)
    try:
        import openpyxl
        from io import BytesIO
        
        print("Parsing as Excel file...")
        wb = openpyxl.load_workbook(BytesIO(content))
        sheet = wb.active
        
        vuids = set()
        
        # Look for VUID column
        headers = [cell.value for cell in sheet[1]]
        vuid_col = None
        
        for idx, header in enumerate(headers):
            if header and 'VUID' in str(header).upper():
                vuid_col = idx
                break
        
        if vuid_col is None:
            # Try to find 10-digit numbers in any column
            print("No VUID column found, scanning all cells...")
            for row in sheet.iter_rows(min_row=2, values_only=True):
                for cell in row:
                    if cell and re.match(r'^\d{10}$', str(cell)):
                        vuids.add(str(cell))
        else:
            # Extract from VUID column
            print(f"Found VUID column at index {vuid_col}")
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[vuid_col]:
                    vuid = str(row[vuid_col]).strip()
                    if re.match(r'^\d{10}$', vuid):
                        vuids.add(vuid)
        
        print(f"Extracted {len(vuids)} unique VUIDs from Excel")
        return list(vuids)
        
    except ImportError:
        print("openpyxl not available, trying pandas...")
        try:
            import pandas as pd
            from io import BytesIO
            
            df = pd.read_excel(BytesIO(content))
            vuids = set()
            
            # Look for VUID column
            vuid_col = None
            for col in df.columns:
                if 'VUID' in str(col).upper():
                    vuid_col = col
                    break
            
            if vuid_col:
                for vuid in df[vuid_col]:
                    if pd.notna(vuid) and re.match(r'^\d{10}$', str(vuid)):
                        vuids.add(str(vuid))
            else:
                # Scan all columns
                for col in df.columns:
                    for val in df[col]:
                        if pd.notna(val) and re.match(r'^\d{10}$', str(val)):
                            vuids.add(str(val))
            
            print(f"Extracted {len(vuids)} unique VUIDs from Excel")
            return list(vuids)
            
        except ImportError:
            print("ERROR: Neither openpyxl nor pandas is installed")
            print("Install with: pip install openpyxl or pip install pandas openpyxl")
            return []
    except Exception as e:
        print(f"Error parsing Excel: {e}")
        return []

def import_to_database(vuids):
    """Import voters to database by cross-referencing VUIDs."""
    return _import_vuids(vuids, 'early-voting', 'hidalgo-roster')

def import_to_database_mailin(vuids):
    """Import mail-in voters."""
    return _import_vuids(vuids, 'mail-in', 'hidalgo-mailin-roster')

def _import_vuids(vuids, voting_method, data_source):
    """Import voters to database by cross-referencing VUIDs."""
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    
    print(f"\nCross-referencing {len(vuids)} VUIDs with voter database...")
    
    # Check which VUIDs exist in our database
    placeholders = ','.join('?' * len(vuids))
    cur.execute(f"""
        SELECT vuid, county, precinct, lat, lng
        FROM voters
        WHERE vuid IN ({placeholders})
    """, vuids)
    
    existing_voters = cur.fetchall()
    print(f"Found {len(existing_voters)} voters in database")
    
    if len(existing_voters) == 0:
        print("\n⚠ No matching voters found in database")
        print("This could mean:")
        print("  1. The VUIDs in the roster don't match the database format")
        print("  2. These voters haven't been imported to the database yet")
        print("  3. The PDF parsing didn't extract VUIDs correctly")
        
        # Show sample VUIDs from roster
        print(f"\nSample VUIDs from roster: {vuids[:5]}")
        
        # Show sample VUIDs from database
        cur.execute("SELECT vuid FROM voters LIMIT 5")
        db_sample = [row[0] for row in cur.fetchall()]
        print(f"Sample VUIDs from database: {db_sample}")
        
        conn.close()
        return 0
    
    imported = 0
    skipped = 0
    missing_location = 0
    
    for row in existing_voters:
        vuid = row[0]
        county = row[1]
        precinct = row[2]
        lat = row[3]
        lng = row[4]
        
        # Check if already imported
        cur.execute("""
            SELECT 1 FROM voter_elections
            WHERE vuid = ? AND election_date = ?
        """, (vuid, ELECTION_DATE))
        
        if cur.fetchone():
            skipped += 1
            continue
        
        # Track voters without location data
        if not lat or not lng:
            missing_location += 1
        
        # Insert new record
        cur.execute("""
            INSERT INTO voter_elections (vuid, election_date, voting_method, data_source, created_at)
            VALUES (?, ?, ?, ?, datetime('now'))
        """, (vuid, ELECTION_DATE, voting_method, data_source))
        
        imported += 1
    
    conn.commit()
    conn.close()
    
    print(f"\n✓ Imported: {imported}")
    print(f"  Skipped (duplicates): {skipped}")
    if missing_location > 0:
        print(f"  ⚠ Missing location data: {missing_location} (won't appear on map)")
    
    return imported

def main():
    print("="*80)
    print("McAllen ISD Bond 2026 - Roster Import")
    print("="*80)
    
    try:
        content = download_roster(ROSTER_URL)
        vuids = parse_roster_pdf(content)
        
        if not vuids:
            print("\n✗ No VUIDs extracted from PDF")
            return 1
        
        imported = import_to_database(vuids)
        
        # Also import mail-in ballots
        print("\nChecking mail-in ballots...")
        try:
            MAILIN_URL = 'https://www.hidalgocounty.us/DocumentCenter/View/72496/ABBM-LIST-May-2-2026-Cumulative'
            mailin_content = download_roster(MAILIN_URL)
            mailin_vuids = parse_roster_pdf(mailin_content)
            if mailin_vuids:
                print(f"Found {len(mailin_vuids)} mail-in VUIDs")
                mailin_imported = import_to_database_mailin(mailin_vuids)
                imported += mailin_imported
        except Exception as e:
            print(f"Mail-in import skipped: {e}")
        
        # Always regenerate ALL caches (not just voters)
        print("\nRegenerating all caches...")
        import subprocess
        try:
            subprocess.run([
                sys.executable,
                '/opt/whovoted/deploy/refresh_bond_caches.py'
            ], check=True)
            print("✓ All caches regenerated")
        except Exception as e:
            print(f"⚠ Cache regeneration failed: {e}")
        
        print("\n" + "="*80)
        if imported > 0:
            print(f"✓ SUCCESS: Imported {imported} voters")
        else:
            print("No new voters to import (all up to date)")
        print(f"Total VUIDs in roster: {len(vuids)}")
        print("View at: https://politiquera.com/misdbond2026/")
        print("="*80)
        
        return 0
        
    except Exception as e:
        print(f"\n✗ Error: {e}")
        import traceback
        traceback.print_exc()
        return 1

if __name__ == '__main__':
    sys.exit(main())

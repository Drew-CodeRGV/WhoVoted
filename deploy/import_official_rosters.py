#!/usr/bin/env python3
"""
Import official roster files directly into the database.
Bypasses the admin panel to ensure 100% accuracy.
"""
import sqlite3
import openpyxl
from pathlib import Path
from datetime import datetime
import sys

DB_PATH = '/opt/whovoted/data/whovoted.db'
ROSTER_DIR = Path('/tmp/official_rosters')

ELECTION_DATE = '2026-03-03'
ELECTION_YEAR = '2026'
ELECTION_TYPE = 'primary'

def process_roster_file(filepath, party, voting_method):
    """Extract VUIDs from a roster file."""
    print(f"\nProcessing {filepath.name}...")
    
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        
        # Find VUID column
        header = [cell.value for cell in ws[1]]
        vuid_col = None
        for idx, col_name in enumerate(header):
            if col_name and 'VUID' in str(col_name).upper():
                vuid_col = idx
                break
        
        if vuid_col is None:
            print(f"  ✗ No VUID column found")
            return []
        
        vuids = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            vuid = row[vuid_col]
            if vuid:
                vuids.append({
                    'vuid': str(vuid).strip(),
                    'party': party,
                    'method': voting_method
                })
        
        wb.close()
        print(f"  ✓ Extracted {len(vuids):,} VUIDs")
        return vuids
        
    except Exception as e:
        print(f"  ✗ Error: {e}")
        return []

def import_to_database(all_records):
    """Import records to database, handling duplicates correctly."""
    print("\n" + "="*70)
    print("IMPORTING TO DATABASE")
    print("="*70)
    
    conn = sqlite3.connect(DB_PATH, timeout=60)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    
    # First, delete existing 2026-03-03 records for Hidalgo County
    print("\nClearing existing 2026-03-03 records for Hidalgo County...")
    deleted = conn.execute("""
        DELETE FROM voter_elections
        WHERE election_date = ?
          AND vuid IN (SELECT vuid FROM voters WHERE county = 'Hidalgo')
    """, [ELECTION_DATE]).rowcount
    conn.commit()
    print(f"  ✓ Deleted {deleted:,} existing records")
    
    # Deduplicate by VUID (keep first occurrence)
    print("\nDeduplicating VUIDs...")
    seen_vuids = {}
    unique_records = []
    
    for record in all_records:
        vuid = record['vuid']
        if vuid not in seen_vuids:
            seen_vuids[vuid] = record
            unique_records.append(record)
    
    duplicates = len(all_records) - len(unique_records)
    print(f"  Total records: {len(all_records):,}")
    print(f"  Unique VUIDs: {len(unique_records):,}")
    print(f"  Duplicates removed: {duplicates:,}")
    
    # Count by party
    dem_count = sum(1 for r in unique_records if r['party'] == 'Democratic')
    rep_count = sum(1 for r in unique_records if r['party'] == 'Republican')
    
    print(f"\n  Democratic: {dem_count:,}")
    print(f"  Republican: {rep_count:,}")
    print(f"  Total: {len(unique_records):,}")
    
    # Insert records
    print("\nInserting records...")
    inserted = 0
    batch_size = 1000
    
    for i in range(0, len(unique_records), batch_size):
        batch = unique_records[i:i+batch_size]
        
        # Insert into voter_elections
        conn.executemany("""
            INSERT OR REPLACE INTO voter_elections 
            (vuid, election_date, election_year, election_type, voting_method, party_voted)
            VALUES (?, ?, ?, ?, ?, ?)
        """, [(r['vuid'], ELECTION_DATE, ELECTION_YEAR, ELECTION_TYPE, r['method'], r['party']) 
              for r in batch])
        
        inserted += len(batch)
        if (i + batch_size) % 10000 == 0:
            print(f"  Progress: {inserted:,} / {len(unique_records):,}")
    
    conn.commit()
    print(f"  ✓ Inserted {inserted:,} records")
    
    # Verify final counts
    print("\n" + "="*70)
    print("VERIFICATION")
    print("="*70)
    
    result = conn.execute("""
        SELECT 
            COUNT(DISTINCT ve.vuid) as unique_vuids,
            SUM(CASE WHEN ve.party_voted = 'Democratic' THEN 1 ELSE 0 END) as dem,
            SUM(CASE WHEN ve.party_voted = 'Republican' THEN 1 ELSE 0 END) as rep
        FROM voter_elections ve
        JOIN voters v ON ve.vuid = v.vuid
        WHERE v.county = 'Hidalgo'
          AND ve.election_date = ?
    """, [ELECTION_DATE]).fetchone()
    
    print(f"\nDatabase (after import):")
    print(f"  Democratic: {result[1]:,}")
    print(f"  Republican: {result[2]:,}")
    print(f"  Total: {result[0]:,}")
    
    print(f"\nOfficial totals:")
    print(f"  Democratic: 49,664")
    print(f"  Republican: 13,217")
    print(f"  Total: 62,881")
    
    print(f"\nDifference:")
    print(f"  Democratic: {result[1] - 49664:+,}")
    print(f"  Republican: {result[2] - 13217:+,}")
    print(f"  Total: {result[0] - 62881:+,}")
    
    conn.close()

# Main execution
print("="*70)
print("IMPORTING OFFICIAL HIDALGO COUNTY ROSTERS")
print("="*70)

files_to_process = [
    ('EV_DEM_Roster_March_3_2026_Cumulative.xlsx', 'Democratic', 'early-voting'),
    ('EV_REP_Roster_March_3_2026_Cumulative.xlsx', 'Republican', 'early-voting'),
    ('ABBM_LIST_March_3_2026_Cumulative_DEM.xlsx', 'Democratic', 'mail-in'),
    ('ABBM_LIST_March_3_2026_Cumulative_REP.xlsx', 'Republican', 'mail-in'),
]

all_records = []

for filename, party, method in files_to_process:
    filepath = ROSTER_DIR / filename
    if filepath.exists():
        records = process_roster_file(filepath, party, method)
        all_records.extend(records)
    else:
        print(f"\n✗ File not found: {filepath}")

if all_records:
    import_to_database(all_records)
    print("\n✅ Import complete!")
else:
    print("\n✗ No records to import")
